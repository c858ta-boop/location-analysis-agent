import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Сокращенный анализ локации", layout="wide")

st.title("🚗 ИИ-Агент: Сокращенный анализ локации")
st.write("Сравнение таблиц на листах 'АФ сокр' из двух отчетов с умной бизнес-подсветкой доходов и расходов.")

# Панель настроек в боковой панели
with st.sidebar:
    st.header("⚙️ Настройки структуры")
    target_column = st.text_input("Название столбца со статьями:", value="Статья")
    type_column = st.text_input("Название столбца типа (Доходы/Расходы):", value="Доходы Расходы")
    header_row = st.number_input("Строка с заголовками (в Excel нумерация с 1):", min_value=1, value=4)
    st.caption("ℹ️ Код 1 = Доходы (Рост=Зеленый, Падение=Красный). Код 2 = Расходы (Рост=Красный, Падение=Зеленый).")

# Блок загрузки файлов
col1, col2 = st.columns(2)
with col1:
    file_1 = st.file_uploader("📂 Загрузите файл 1 (Прошлый период / База)", type=["xlsx"])
with col2:
    file_2 = st.file_uploader("📂 Загрузите файл 2 (Текущий период / Отчет)", type=["xlsx"])

def is_colored(cell):
    """Проверяет, есть ли у ячейки цветная заливка"""
    if cell and cell.fill and cell.fill.fill_type:
        color = cell.fill.start_color.index
        if color and str(color) not in ['00000000', '0', 'FFFFFFFF', 'System_Color_Window']:
            return True
    return False

def get_colored_rows(file_bytes, sheet_name, header_idx, target_col_name):
    """Быстро находит строки с цветовой заливкой с защитой от пустых ячеек"""
    colored_rows = set()
    try:
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            target_col_idx = None
            
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_idx, column=col).value
                if cell_val is not None and str(cell_val).strip() == target_col_name:
                    target_col_idx = col
                    break
                    
            if target_col_idx:
                for row_idx in range(header_idx + 1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=target_col_idx)
                    if cell and cell.value is not None and is_colored(cell):
                        colored_rows.add(row_idx - header_idx - 1)
    except:
        pass
    return colored_rows

def clean_to_float(val):
    """Всеядная функция для приведения ячеек к числу с плавающей точкой"""
    if pd.isna(val) or val is None:
        return 0.0
    val_str = str(val).strip()
    if val_str == "" or val_str == "-":
        return 0.0
    try:
        val_str = val_str.replace('\xa0', '').replace(' ', '').replace(',', '.')
        return float(val_str)
    except:
        return 0.0

# Основная логика приложения
if file_1 and file_2:
    st.success("Файлы успешно загружены! Начинаю факторный анализ...")
    
    pandas_header_index = int(header_row) - 1
    sheet_target = "АФ сокр"
    
    # Клонируем файлы в независимые буферы памяти для стабильного чтения
    old_bytes = file_1.read()
    new_bytes = file_2.read()
    
    xl_1 = pd.ExcelFile(BytesIO(old_bytes))
    xl_2 = pd.ExcelFile(BytesIO(new_bytes))
    
    if sheet_target not in xl_1.sheet_names or sheet_target not in xl_2.sheet_names:
        st.error(f"❌ Ошибка: Лист '{sheet_target}' не найден в одном или обоих файлах!")
    else:
        # Загружаем датафреймы
        df_1 = pd.read_excel(BytesIO(old_bytes), sheet_name=sheet_target, header=pandas_header_index)
        df_2 = pd.read_excel(BytesIO(new_bytes), sheet_name=sheet_target, header=pandas_header_index)
        
        df_1.columns = [str(c).strip() for c in df_1.columns]
        df_2.columns = [str(c).strip() for c in df_2.columns]
        
        if target_column not in df_1.columns or target_column not in df_2.columns:
            st.error(f"❌ Столбец '{target_column}' не найден на листе '{sheet_target}'. Проверьте настройки в боковом меню.")
        elif type_column not in df_2.columns:
            st.error(f"❌ Столбец типа '{type_column}' не найден в новом файле. Проверьте заголовки.")
        else:
            # Очищаем от пустых значений и дубликатов
            df_1 = df_1.dropna(subset=[target_column])
            df_2 = df_2.dropna(subset=[target_column])
            df_1[target_column] = df_1[target_column].astype(str).str.strip()
            df_2[target_column] = df_2[target_column].astype(str).str.strip()
            
            # Находим все числовые столбцы для сравнения
            numeric_cols = [col for col in df_2.columns if col != target_column and col != type_column and col in df_1.columns and not str(col).startswith('Unnamed:')]
            
            # Создаем каркас результирующей таблицы на основе Файла 2
            df_result_raw = df_2[[target_column, type_column] + numeric_cols].copy()
            df_result = df_result_raw.copy()
            
            for col in numeric_cols:
                df_result[col] = df_result[col].astype(object)
                
            df_1_indexed = df_1.set_index(target_column)
            
            # Матрица цветов для отображения в Streamlit
            color_matrix = pd.DataFrame('', index=df_result.index, columns=df_result.columns)
            
            # Константы стилей для ячеек
            STYLE_GREEN = 'background-color: #D1FAE5; color: #065F46;' # Хорошо для бизнеса
            STYLE_RED = 'background-color: #FEE2E2; color: #991B1B;'   # Плохо для бизнеса
            
            # Пересчитываем каждую ячейку
            for idx, row in df_result_raw.iterrows():
                statya = row[target_column]
                row_type = str(row[type_column]).strip()
                
                for col in numeric_cols:
                    val_2 = row[col]
                    
                    try:
                        val_1 = df_1_indexed.loc[statya, col]
                        if isinstance(val_1, pd.Series):
                            val_1 = val_1.iloc
                    except KeyError:
                        val_1 = 0.0
                        
                    val_2_float = clean_to_float(val_2)
                    val_1_float = clean_to_float(val_1)
                    
                    delta = val_2_float - val_1_float
                    
                    if delta > 0:
                        df_result.at[idx, col] = f"{val_2_float:,.2f} (+{delta:,.2f})"
                        # ЕСЛИ РОСТ ДОХОДОВ (1) -> ЗЕЛЕНЫЙ. ЕСЛИ РОСТ РАСХОДОВ (2) -> КРАСНЫЙ.
                        color_matrix.at[idx, col] = STYLE_GREEN if row_type == "1" else STYLE_RED
                    elif delta < 0:
                        df_result.at[idx, col] = f"{val_2_float:,.2f} (-{abs(delta):,.2f})"
                        # ЕСЛИ ПАДЕНИЕ ДОХОДОВ (1) -> КРАСНЫЙ. ЕСЛИ ПАДЕНИЕ РАСХОДОВ (2) -> ЗЕЛЕНЫЙ.
                        color_matrix.at[idx, col] = STYLE_RED if row_type == "1" else STYLE_GREEN
                    else:
                        df_result.at[idx, col] = f"{val_2_float:,.2f}"
            
            def style_cells(df):
                return color_matrix
            
            st.subheader("📊 Результаты сравнительного анализа локации")
            st.write("Цветовая индикация адаптирована под экономику ДЦ: рост доходов и падение расходов подсвечены **зеленым**, падение доходов и рост расходов — **красным**.")
            st.dataframe(df_result.style.apply(style_cells, axis=None), use_container_width=True)
            
            # Построение вывода печатной формы прямо на экран
            st.write("---")
            st.subheader("🖨️ Печать и экспорт в PDF")
            st.write("Нажмите комбинацию клавиш **Ctrl + P** (или **Cmd + P** на Mac) прямо на этой странице браузера, чтобы сохранить этот отчет в PDF.")
            
            html_preview = "<div style='font-family: Arial, sans-serif; padding: 20px; border: 1px solid #E5E7EB; border-radius: 5px; background: white;'>\n"
            html_preview += "<h2 style='color: #1E3A8A; border-bottom: 2px solid #1E3A8A; padding-bottom: 8px; font-size: 18px; margin-top:0;'>Сокращенный анализ локации (Бизнес-отчет)</h2>\n"
            html_preview += "<table style='width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 11px;'>\n"
            
            html_preview += "<tr style='background: #1E3A8A; color: white;'>\n"
            html_preview += f"<th style='padding: 6px; text-align: left;'>{target_column}</th>\n"
            html_preview += "<th style='padding: 6px; text-align: center;'>Тип</th>\n"
            for col in numeric_cols:
                html_preview += f"<th style='padding: 6px; text-align: right;'>{col}</th>\n"
            html_preview += "</tr>\n"
            
            for idx, row in df_result.iterrows():
                bg_row = "#F9FAFB" if idx % 2 == 0 else "#FFFFFF"
                r_type = str(row[type_column]).strip()
                type_label = "Доход" if r_type == "1" else "Расход"
                
                html_preview += f"<tr style='background: {bg_row}; border-bottom: 1px solid #E5E7EB;'>\n"
                html_preview += f"<td style='padding: 6px;'><b>{row[target_column]}</b></td>\n"
                html_preview += f"<td style='padding: 6px; text-align: center; color: #6B7280;'>{type_label}</td>\n"
                
                for col in numeric_cols:
                    cell_text = str(row[col])
                    cell_style = "padding: 6px; text-align: right;"
                    
                    if "(+" in cell_text:
                        # Исправлено: РОСТ ДОХОДОВ (1) -> ЗЕЛЕНЫЙ, РОСТ РАСХОДОВ (2) -> КРАСНЫЙ
                        cell_style += " background-color: #D1FAE5; color: #065F46;" if r_type == "1" else " background-color: #FEE2E2; color: #991B1B;"
                    elif "(-" in cell_text:
